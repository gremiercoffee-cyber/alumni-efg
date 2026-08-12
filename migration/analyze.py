"""
Dedupe review pass for the Aish Gesher alumni migration.

Reads the two source workbooks, builds a canonical person list keyed on the
AlumniID already used by the rebbeim connection sheet, and proposes merges for
review. Nothing is written to a database here -- this only produces evidence.

Re-runnable: tweak the matching rules and run again. Decisions you have already
made are read back from decisions.json and preserved.

    python migration/analyze.py

Outputs (in migration/out/):
    people.json             canonical people, enriched from the year tabs
    enrollments.json        person x academic_year x program_level, deduped
    merge_candidates.json   proposed merges with evidence + confidence tier
    review.html            the page you actually look at
"""

import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# --- where things live ------------------------------------------------------

DOWNLOADS = Path.home() / "Downloads"
ALUMNI_WB = DOWNLOADS / "Aish Gesher Alumni Database - UPDATED.xlsx"
REBBEIM_WB = DOWNLOADS / "Alumni x Rebbeim.xlsx"

OUT = Path(__file__).parent / "out"
DECISIONS = Path(__file__).parent / "decisions.json"

# Tabs in the rebbeim workbook that are derived views, not a rebbe's answers.
DERIVED_TABS = {
    "Rebbe → Alumni",
    "Alumni Coverage",
    "Unclaimed Alumni",
    "Staff Network",
    "Alumni",
    "Staff",
}

# Program levels, normalized. Spelling drifted across 14 years of tabs.
PROGRAM_ALIASES = {
    "shana alef": "Shana Alef",
    "shana bet": "Shana Bet",
    "shana gimel": "Shana Gimel",
    "shana gimmel": "Shana Gimel",
    "shana daled": "Shana Daled",
    "madrich": "Madrich",
    "shana alef/bet": "Shana Alef",  # one row; the Bet year is its own enrollment
}

# First-name equivalences. Hebrew/English pairs and the nicknames that actually
# show up in this data -- not a general-purpose list.
NAME_EQUIV = [
    {"joseph", "yosef", "yossi", "joey", "yossef", "jo"},
    {"joshua", "josh", "yehoshua", "shua", "shuey", "shia"},
    {"jacob", "yaakov", "yakov", "jake", "koby"},
    {"emanuel", "manny", "menachem"},
    {"abraham", "avraham", "avrohom", "avi", "abe", "avromi"},
    {"isaac", "issac", "yitzchak", "yitzy", "itzy"},
    {"david", "dovid", "dave", "duvi"},
    {"aaron", "aron", "aharon", "aahron", "arik"},
    {"benjamin", "ben", "benny", "binyamin", "benzion"},
    {"daniel", "dan", "danny", "dani"},
    {"michael", "mike", "micha", "mic", "mikey"},
    {"gabriel", "gabe", "gavriel", "gav"},
    {"samuel", "sam", "shmuel", "shmulie", "shmully", "shmuly"},
    {"nathan", "natan", "nat", "nate"},
    {"eliyahu", "eli", "elijah", "elie"},
    {"matthew", "mathew", "matt", "mati"},
    {"alexander", "alex", "sasha"},
    {"reuben", "ruben", "reuven"},
    {"solomon", "shlomo", "sol"},
    {"zev", "zevi", "wolf"},
    {"moshe", "moe", "mo", "morris"},
    {"kalman", "klonymus"},
    {"ezekial", "ezekiel", "zeke", "yechezkel"},
]


# --- text helpers -----------------------------------------------------------


def txt(v):
    """A cell as a clean string. openpyxl hands back floats, None, whitespace."""
    if v is None:
        return ""
    s = str(v).strip()
    # Zip codes and phone numbers arrive as '10901.0'
    if re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return s


def norm(s):
    """Lowercase, strip accents and everything that isn't a letter."""
    s = unicodedata.normalize("NFKD", txt(s))
    return re.sub(r"[^a-z]", "", s.lower())


def split_name(full):
    """'Joshua (Shua) Keslassy' -> ('joshua', 'keslassy', 'shua')"""
    full = txt(full).replace(",", " ")
    paren = re.search(r"\(([^)]*)\)", full)
    nickname = norm(paren.group(1)) if paren else ""
    full = re.sub(r"\([^)]*\)", " ", full)
    parts = [p for p in full.split() if p.strip()]
    if not parts:
        return "", "", nickname
    if len(parts) == 1:
        return norm(parts[0]), "", nickname
    return norm(parts[0]), norm(" ".join(parts[1:])), nickname


def lev(a, b):
    """Levenshtein distance. Short strings only, so the naive version is fine."""
    if a == b:
        return 0
    if not a or not b:
        return max(len(a), len(b))
    prev = list(range(len(a) + 1))
    for i, cb in enumerate(b, 1):
        cur = [i]
        for j, ca in enumerate(a, 1):
            cur.append(min(prev[j - 1] + (ca != cb), cur[j - 1] + 1, prev[j] + 1))
        prev = cur
    return prev[-1]


def first_names_equivalent(a, b):
    """Are these two first names plausibly the same person?"""
    if not a or not b:
        return False
    if a == b:
        return True
    for group in NAME_EQUIV:
        if a in group and b in group:
            return True
    # 'Shmuel' vs 'Shmulie' -- one is a prefix of the other
    if len(a) >= 3 and len(b) >= 3 and (a.startswith(b) or b.startswith(a)):
        return True
    return False


def clean_email(v):
    e = txt(v).lower()
    return e if "@" in e and " " not in e else ""


def clean_program(v):
    return PROGRAM_ALIASES.get(txt(v).lower(), "")


# --- loading ----------------------------------------------------------------


def load_year_tabs(wb):
    """
    Every enrollment row across the 14 year tabs.

    Years 1-12 share one column layout. Years 13-14 are a different, intake-style
    sheet: parent contacts and high school, no address/college/rebbe. Year 13
    keeps the program level in an unlabeled third column; year 14 has none.
    """
    rows = []
    for name in wb.sheetnames:
        m = re.match(r"^(\d+)(?:st|nd|rd|th) Year - (\d{4})-(\d{4})", name)
        if not m:
            continue
        year_no, academic_year = int(m.group(1)), f"{m.group(2)}-{m.group(3)}"
        data = list(wb[name].iter_rows(values_only=True))
        if not data:
            continue
        header = [txt(c) for c in data[0]]

        for raw in data[1:]:
            r = list(raw) + [None] * 30
            if year_no == 13:
                rec = dict(
                    first=txt(r[0]), last=txt(r[1]),
                    program=clean_program(r[2]), email=clean_email(r[3]),
                    phone=txt(r[5]), city=txt(r[11]), state=txt(r[12]),
                    high_school=txt(r[13]),
                    parents=[
                        dict(relation="father", name=txt(r[4]), phone=txt(r[6]), email=clean_email(r[7])),
                        dict(relation="mother", name=txt(r[8]), phone=txt(r[9]), email=clean_email(r[10])),
                    ],
                )
            elif year_no == 14:
                rec = dict(
                    first=txt(r[1]), last=txt(r[0]),
                    # The 2025-2026 tab has no program column; the user confirmed
                    # this whole intake is Shana Alef.
                    program="Shana Alef", email=clean_email(r[3]),
                    phone=txt(r[4]), city=txt(r[11]), state=txt(r[12]),
                    high_school=txt(r[13]),
                    parents=[
                        dict(relation="father", name=txt(r[5]), phone=txt(r[7]), email=clean_email(r[6])),
                        dict(relation="mother", name=txt(r[8]), phone=txt(r[10]), email=clean_email(r[9])),
                    ],
                )
            else:
                # 'Progam' is a real typo in the 7th Year tab header.
                rec = dict(
                    first=txt(r[1]), last=txt(r[0]),
                    program=clean_program(r[3]), email=clean_email(r[4]),
                    phone=txt(r[5]), street=txt(r[6]), city=txt(r[7]),
                    state=txt(r[8]), zip=txt(r[9]), country=txt(r[10]),
                    college=txt(r[11]), grad_school=txt(r[12]),
                    occupation=txt(r[13]), spouse=txt(r[14]),
                    contact_updated=txt(r[15]), rebbe=txt(r[16]),
                    notes=txt(r[17]), spotlight=txt(r[18]),
                    learning_post=txt(r[19]), aish_impact=txt(r[20]),
                    parents=[],
                )

            if not (rec["first"] or rec["last"]):
                continue
            # Stats rows leaked into some year tabs ('13.0 Total Shana Alef').
            if re.fullmatch(r"[\d.]+", rec["first"]) or "total" in rec["last"].lower():
                continue
            rec["academic_year"] = academic_year
            rec["source_tab"] = name
            rec["parents"] = [p for p in rec.get("parents", []) if p["name"] or p["phone"] or p["email"]]
            rows.append(rec)
    return rows


def load_canonical(wb):
    """The Alumni tab -- 743 people with the AlumniIDs the rebbeim sheet uses."""
    people = []
    for r in list(wb["Alumni"].iter_rows(values_only=True))[1:]:
        name = txt(r[1])
        if not name:
            continue
        aid = txt(r[0])
        years = [y.strip() for y in txt(r[3]).split(",") if y.strip()]
        first, last, nick = split_name(name)
        people.append(dict(
            alumni_id=int(float(aid)) if aid else None,
            name=name, first=first, last=last, nickname=nick,
            # deduped: the source lists '13th Year - 2024-2025' twice for 115 people
            years=sorted(set(years)),
            years_raw_count=len(years),
        ))
    return people


def load_touchpoints(wb):
    """
    The 'Rebbe Contact - Alumni 2022' tab is a campaign log: one column per
    outreach push (Summer 5782, Chanukah 5783, Pesach 5783), TRUE if that alumnus
    was reached. Becomes interaction rows.
    """
    name = "Rebbe Contact - Alumni 2022"
    if name not in wb.sheetnames:
        return []
    rows = list(wb[name].iter_rows(values_only=True))
    header = [txt(c) for c in rows[0]]
    campaigns = [(i, h) for i, h in enumerate(header) if i >= 4 and h]
    out = []
    for r in rows[1:]:
        r = list(r) + [None] * 12
        last, first = txt(r[0]), txt(r[1])
        if not (first or last):
            continue
        for i, campaign in campaigns:
            if txt(r[i]).lower() == "true":
                out.append(dict(first=first, last=last, campaign=campaign,
                                rebbe=txt(r[3])))
    return out


def load_staff(wb):
    """
    All 30 rabbis, from the connection workbook -- one tab each, plus a Staff tab.

    This is the authoritative staff list. The alumni database's 'Rebbe Contact'
    column only ever names about 15 of them, and uses a different convention
    ("R' Rosman" there vs "Rav Rosman" here), so we map those strings on rather
    than treating them as a second set of people.
    """
    tabs = [n for n in wb.sheetnames if n not in DERIVED_TABS]
    listed = [txt(r[0]) for r in wb["Staff"].iter_rows(values_only=True) if txt(r[0])]

    staff = []
    for name in sorted(set(tabs) | set(listed)):
        m = re.match(r"^(Rabbi|Rav|R'|R)\s+(.*)$", name)
        staff.append(dict(
            name=name,
            title=m.group(1) if m else "",
            surname=m.group(2).strip() if m else name,
            has_own_tab=name in tabs,
            in_staff_tab=name in listed,
        ))
    return staff


def map_rebbe_strings(raw_values, staff):
    """
    Fold the alumni database's free-text 'Rebbe Contact' values onto the staff
    list. Matches on surname, since only the title differs. Anything that does
    not match is reported rather than silently dropped.
    """
    by_surname = {norm(s["surname"]): s["name"] for s in staff}
    mapping, unmapped = {}, Counter()
    for raw, count in raw_values.items():
        if not raw:
            continue
        # A few cells hold two rabbis ("R' Hershman/R' Mintz") or junk ("??").
        parts = [p for p in re.split(r"[/,&]| and ", raw) if p.strip()]
        hits = []
        for part in parts:
            m = re.match(r"^(?:Rabbi|Rav|R'|R)?\s*(.*)$", part.strip())
            key = norm(m.group(1) if m else part)
            if not key:
                continue
            if key in by_surname:
                hits.append(by_surname[key])
                continue
            # "R' Wolf" vs "Rabbi Wolff", "R' Greene" vs "Rabbi Green"
            near = [v for k, v in by_surname.items() if len(k) > 3 and lev(k, key) <= 1]
            if len(near) == 1:
                hits.append(near[0])
        if hits:
            mapping[raw] = sorted(set(hits))
        else:
            unmapped[raw] = count
    return mapping, unmapped


def add_missing_staff(staff, unmapped):
    """
    Rabbis who appear in the alumni database's 'Rebbe Contact' column but have no
    tab in the connection workbook. They are real staff with real alumni attached
    -- they just were not included when that sheet was built, so they carry
    through to the schema with no connections recorded yet.
    """
    added = []
    for raw, count in unmapped.items():
        m = re.match(r"^(Rabbi|Rav|R'|R)\s+([A-Za-z][A-Za-z' -]*)$", raw.strip())
        if not m:
            continue  # '??', a stray date, a rebbe's name typed into the wrong column
        surname = m.group(2).strip()
        name = f"Rabbi {surname}"
        if any(norm(s["surname"]) == norm(surname) for s in staff):
            continue
        staff.append(dict(name=name, title="Rabbi", surname=surname,
                          has_own_tab=False, in_staff_tab=False,
                          source="alumni database only", alumni_referencing=count))
        added.append((name, count))
    staff.sort(key=lambda s: s["surname"])
    return added


DNC_PATTERNS = [
    "do not contact", "don't contact", "dont contact", "not be contacted",
    "not to be contacted", "do not reach out", "asked to leave", "asked not to",
]

# Where each person-level field comes from in a year row. The most recent year
# that has a value wins, so a 2025 address beats the 2019 one.
PERSON_FIELDS = [
    ("phone", "phone"), ("street_address", "street"), ("city", "city"),
    ("state", "state"), ("zip_code", "zip"), ("country", "country"),
    ("college", "college"), ("grad_school", "grad_school"),
    ("occupation", "occupation"), ("spouse_name", "spouse"),
    ("high_school", "high_school"), ("notes", "notes"),
    ("learning_post_gesher", "learning_post"), ("aish_impact", "aish_impact"),
]

COUNTRY_FIX = {"ca": "Canada", "usa": "USA", "us": "USA", "uk": "UK",
               "england": "UK", "israel": "Israel"}


def parse_date(s):
    """The sheets hold 13 different date shapes. Return ISO, or None."""
    s = txt(s)
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)          # '2022-01-03 00:00:00'
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{1,2})/+(\d{1,2})/(\d{2,4})$", s)   # '1/3/22', '5//21/23'
    if m:
        mo, d, y = (int(x) for x in m.groups())
        y = y + 2000 if y < 100 else y
        if 1 <= mo <= 12 and 1 <= d <= 31 and 2000 <= y <= 2100:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def consolidate(people, enroll_by_person, rebbe_map):
    """
    Collapse each person's year rows into one profile.

    A man who was here four years has four rows, each with its own snapshot of
    his address and occupation. The most recent non-empty value wins, which is
    almost always the right answer for contact details.
    """
    for p in people:
        rows = sorted(enroll_by_person.get(p["alumni_id"], []),
                      key=lambda r: r["academic_year"], reverse=True)

        for field, src in PERSON_FIELDS:
            for r in rows:
                if txt(r.get(src)):
                    p[field] = txt(r[src])
                    break

        if p.get("country"):
            p["country"] = COUNTRY_FIX.get(p["country"].lower().strip(), p["country"])
            # One row had a postal code in the country column.
            if re.fullmatch(r"[A-Z0-9 ]{5,8}", p["country"]) and not p["country"].isalpha():
                p.pop("country")

        for r in rows:
            d = parse_date(r.get("contact_updated"))
            if d:
                p["contact_updated_on"] = d
                break

        p["spotlight"] = any(txt(r.get("spotlight")).lower() == "true" for r in rows)

        # 'do not contact' was buried in free text, and in one row in the date
        # column. Anything automated would have contacted these men regardless.
        haystack = " ".join(
            txt(r.get(k)) for r in rows for k in ("notes", "contact_updated", "rebbe")
        ).lower()
        hit = next((pat for pat in DNC_PATTERNS if pat in haystack), None)
        if hit:
            p["do_not_contact"] = True
            p["do_not_contact_reason"] = next(
                (txt(r["notes"]) for r in rows if txt(r.get("notes"))), hit)

        if rows:
            last = max(int(r["academic_year"].split("-")[1]) for r in rows)
            p["expected_graduation_year"] = last + 4

        p["family_contacts"] = [
            dict(relation=c["relation"], name=txt(c["name"]),
                 email=c["email"], phone=txt(c["phone"]))
            for r in rows for c in r.get("parents", [])
        ][:2]

        # Prefer an email that still looks live over one from a decade ago.
        if p.get("emails"):
            p["email"] = p["emails"][0]
            for r in rows:
                if r.get("email"):
                    p["email"] = r["email"]
                    break


def apply_decisions(people, decisions):
    """
    Fold approved merges into the canonical list.

    The lower AlumniID survives, so the key the rebbeim are filling in stays valid.
    The losing record's name is kept as an alias -- that is what lets year-tab rows
    written under the other spelling ("Jason Heideman") attach to the surviving
    person. Losing IDs are recorded so connections can be remapped rather than lost.
    """
    survivor_of = {}

    def resolve(pid):
        while pid in survivor_of:
            pid = survivor_of[pid]
        return pid

    for key, verdict in decisions.items():
        if verdict != "merge":
            continue
        a, b = (int(x) for x in key.split("-"))
        a, b = resolve(a), resolve(b)
        if a == b:
            continue
        keep, drop = min(a, b), max(a, b)
        survivor_of[drop] = keep

    by_id = {p["alumni_id"]: p for p in people}
    for drop, _ in sorted(survivor_of.items()):
        loser, winner = by_id.get(drop), by_id.get(resolve(drop))
        if not loser or not winner or loser is winner:
            continue
        winner.setdefault("aliases", [])
        for n in [loser["name"]] + loser.get("aliases", []):
            if n != winner["name"] and n not in winner["aliases"]:
                winner["aliases"].append(n)
        winner.setdefault("merged_ids", []).append(loser["alumni_id"])
        for e in loser.get("emails", []):
            winner.setdefault("emails", [])
            if e not in winner["emails"]:
                winner["emails"].append(e)
        winner["years"] = sorted(set(winner.get("years", [])) | set(loser.get("years", [])))

    kept = [p for p in people if p["alumni_id"] not in survivor_of]
    return kept, {d: resolve(d) for d in survivor_of}


def load_connections(wb):
    """Which rabbis marked which alumni. Keyed on AlumniID."""
    by_person = defaultdict(list)
    per_rebbe = {}
    for name in wb.sheetnames:
        if name in DERIVED_TABS:
            continue
        n = 0
        for r in list(wb[name].iter_rows(values_only=True))[1:]:
            if txt(r[2]).lower() == "true":
                aid = txt(r[0])
                if aid:
                    by_person[int(float(aid))].append(name)
                    n += 1
        per_rebbe[name] = n
    return by_person, per_rebbe


# --- matching ---------------------------------------------------------------


def suggest_for_unmatched(unmatched, people, conns):
    """
    Year-tab rows whose name matches nobody in the canonical list.

    These are not duplicates of each other -- they are a single spelling that only
    ever appeared in a year tab ("Jason Heideman" where the canonical list has
    "Jason Heidman"). For each, offer the closest canonical people so the row can
    be attached to an existing person or admitted as a new one.
    """
    out = []
    for row in unmatched:
        f, l, nick = split_name(f"{row['first']} {row['last']}")
        scored = []
        for p in people:
            ld = lev(l, p["last"])
            if ld > 2:
                continue
            fd = lev(f, p["first"])
            equiv = first_names_equivalent(f, p["first"]) or (
                nick and first_names_equivalent(nick, p["first"]))
            same_email = row["email"] and row["email"] in p.get("emails", [])
            if same_email:
                score = 0
            elif ld == 0 and equiv:
                score = 1
            elif ld <= 1 and equiv:
                score = 2
            elif ld <= 1 and fd <= 2:
                score = 3
            elif ld == 2 and (equiv or fd == 0):
                score = 4
            elif ld == 0 and f[:1] == p["first"][:1]:
                score = 5
            else:
                continue
            scored.append((score, ld + fd, p))
        scored.sort(key=lambda t: (t[0], t[1]))
        out.append(dict(
            key=f"u{len(out)}",
            row=dict(name=f"{row['first']} {row['last']}".strip(),
                     year=row["academic_year"], program=row["program"],
                     email=row["email"], tab=row["source_tab"]),
            suggestions=[dict(id=p["alumni_id"], name=p["name"],
                              years=p.get("years", []), emails=p.get("emails", []),
                              rebbeim=len(conns.get(p["alumni_id"], [])),
                              certain=(s == 0))
                         for s, _, p in scored[:3]],
        ))
    return out


def build_candidates(people, enroll_by_person, conns):
    """
    Propose merges, with the evidence for each so a human can judge.

    Three tiers:
      certain  identical name, or a shared email address
      likely   same-ish last name plus an equivalent/variant first name
      review   weaker signal -- expect false positives here
    """
    cands = []
    n = len(people)
    for i in range(n):
        a = people[i]
        for j in range(i + 1, n):
            b = people[j]
            if not (a["last"] and b["last"]):
                continue

            last_d = lev(a["last"], b["last"])
            first_d = lev(a["first"], b["first"])
            if last_d > 2:
                continue

            emails_a = {e for e in a.get("emails", []) if e}
            emails_b = {e for e in b.get("emails", []) if e}
            shared_email = sorted(emails_a & emails_b)

            equiv = first_names_equivalent(a["first"], b["first"])
            nick_hit = bool(
                (a["nickname"] and first_names_equivalent(a["nickname"], b["first"]))
                or (b["nickname"] and first_names_equivalent(b["nickname"], a["first"]))
            )

            if shared_email:
                tier, why = "certain", "shares an email address"
            elif last_d == 0 and first_d == 0:
                tier, why = "certain", "identical name"
            elif last_d == 0 and (equiv or nick_hit):
                tier, why = "likely", "same last name, equivalent first name"
            elif last_d <= 1 and (equiv or nick_hit):
                tier, why = "likely", "last name off by one letter, equivalent first name"
            elif last_d <= 1 and first_d <= 2 and a["first"][:1] == b["first"][:1]:
                tier, why = "review", "similar last name, similar first name"
            elif last_d == 0 and a["first"][:1] == b["first"][:1]:
                tier, why = "review", "same last name, same first initial"
            else:
                continue

            ya = {y["academic_year"] for y in enroll_by_person.get(a["alumni_id"], [])}
            yb = {y["academic_year"] for y in enroll_by_person.get(b["alumni_id"], [])}
            overlap = sorted(ya & yb)

            # Two people in the same year with near-identical names are usually a
            # duplicate; but it is also exactly how real brothers/cousins look.
            cands.append(dict(
                key=f"{a['alumni_id']}-{b['alumni_id']}",
                tier=tier, reason=why,
                a=dict(id=a["alumni_id"], name=a["name"], years=sorted(ya) or a["years"],
                       emails=sorted(emails_a), rebbeim=conns.get(a["alumni_id"], [])),
                b=dict(id=b["alumni_id"], name=b["name"], years=sorted(yb) or b["years"],
                       emails=sorted(emails_b), rebbeim=conns.get(b["alumni_id"], [])),
                same_year=overlap,
                connections_at_stake=len(conns.get(a["alumni_id"], [])) + len(conns.get(b["alumni_id"], [])),
            ))

    order = {"certain": 0, "likely": 1, "review": 2}
    cands.sort(key=lambda c: (order[c["tier"]], -c["connections_at_stake"], c["a"]["name"]))
    return cands


# --- main -------------------------------------------------------------------


def main():
    for p in (ALUMNI_WB, REBBEIM_WB):
        if not p.exists():
            sys.exit(f"Missing workbook: {p}")

    OUT.mkdir(parents=True, exist_ok=True)

    awb = openpyxl.load_workbook(ALUMNI_WB, data_only=True)
    rwb = openpyxl.load_workbook(REBBEIM_WB, data_only=True)

    year_rows = load_year_tabs(awb)
    people = load_canonical(rwb)
    staff = load_staff(rwb)
    conns, per_rebbe = load_connections(rwb)

    rebbe_strings = Counter(r.get("rebbe", "") for r in year_rows if r.get("rebbe"))
    rebbe_map, rebbe_unmapped = map_rebbe_strings(rebbe_strings, staff)
    newly_added = add_missing_staff(staff, rebbe_unmapped)
    if newly_added:
        rebbe_map2, rebbe_unmapped = map_rebbe_strings(rebbe_strings, staff)
        rebbe_map.update(rebbe_map2)

    prior = json.loads(DECISIONS.read_text()) if DECISIONS.exists() else {}
    people, merged_into = apply_decisions(people, prior)

    # Remap the rebbeim's answers onto surviving records, then drop the duplicates
    # a merge creates (two rabbis marking what turned out to be one person).
    remapped = defaultdict(set)
    for pid, rebbeim in conns.items():
        remapped[merged_into.get(pid, pid)].update(rebbeim)
    conns = {pid: sorted(r) for pid, r in remapped.items()}

    # Attach each year-tab row to a canonical person by name, aliases included.
    index = defaultdict(list)
    for p in people:
        index[(p["first"], p["last"])].append(p)
        if p["nickname"]:
            index[(p["nickname"], p["last"])].append(p)
        for alias in p.get("aliases", []):
            af, al, an = split_name(alias)
            index[(af, al)].append(p)
            if an:
                index[(an, al)].append(p)

    by_id = {p["alumni_id"]: p for p in people}
    next_id = max(by_id) + 1
    enroll_by_person = defaultdict(list)
    unmatched = []
    for row in year_rows:
        f, l, nick = split_name(f"{row['first']} {row['last']}")
        # An explicit decision always wins over name matching. Two different people
        # can share a name exactly (there are two unrelated Avi Greens), and matching
        # on the name alone silently folds one into the other.
        verdict = prior.get(f"row:{row['first']} {row['last']}".strip() + f"|{row['source_tab']}")
        hit = None
        if verdict and verdict != "new":
            target = by_id.get(merged_into.get(int(verdict), int(verdict)))
            if target:
                # Keep the spelling this row used, so the man is still findable
                # under the name whoever typed it knew him by.
                alias = f"{row['first']} {row['last']}".strip()
                target.setdefault("aliases", [])
                if alias != target["name"] and alias not in target["aliases"]:
                    target["aliases"].append(alias)
                    af, al, _ = split_name(alias)
                    index[(af, al)].append(target)
                hit = [target]
        if not hit and not verdict:
            hit = index.get((f, l)) or index.get((nick, l)) if nick else index.get((f, l))
            if not hit:
                # fall back to a first-name-equivalent match on an exact last name
                hit = [p for p in people if p["last"] == l and first_names_equivalent(p["first"], f)]
        if not hit:
            # An answered row from a previous review round. Keyed on content rather
            # than position so re-running the matcher cannot shift an answer onto a
            # different row.
            verdict = prior.get(f"row:{row['first']} {row['last']}".strip() + f"|{row['source_tab']}")
            if verdict == "new":
                p = dict(alumni_id=next_id, name=f"{row['first']} {row['last']}".strip(),
                         first=f, last=l, nickname=nick, years=[], years_raw_count=0,
                         emails=[e for e in [row["email"]] if e], origin="added during review")
                next_id += 1
                people.append(p)
                by_id[p["alumni_id"]] = p
                index[(f, l)].append(p)
                hit = [p]
            elif verdict:
                target = by_id.get(merged_into.get(int(verdict), int(verdict)))
                if target:
                    target.setdefault("aliases", [])
                    alias = f"{row['first']} {row['last']}".strip()
                    if alias != target["name"] and alias not in target["aliases"]:
                        target["aliases"].append(alias)
                        af, al, an = split_name(alias)
                        index[(af, al)].append(target)
                    hit = [target]
        if hit:
            person = hit[0]
            enroll_by_person[person["alumni_id"]].append(row)
            person.setdefault("emails", [])
            if row["email"] and row["email"] not in person["emails"]:
                person["emails"].append(row["email"])
            person.setdefault("sources", []).append(row["source_tab"])
        else:
            unmatched.append(row)

    for p in people:
        p.setdefault("emails", [])

    consolidate(people, enroll_by_person, rebbe_map)

    # Enrollments, deduped -- this is the fix for the 115 doubled year entries.
    enrollments, seen = [], set()
    for pid, rows in enroll_by_person.items():
        for r in rows:
            k = (pid, r["academic_year"], r["program"])
            if k in seen:
                continue
            seen.add(k)
            enrollments.append(dict(person_id=pid, academic_year=r["academic_year"],
                                    program_level=r["program"], rebbe=r.get("rebbe", "")))

    cands = build_candidates(people, enroll_by_person, conns)
    orphans = suggest_for_unmatched(unmatched, people, conns)
    (OUT / "unmatched.json").write_text(json.dumps(orphans, indent=1))

    # Resolve each enrollment's free-text rebbe onto a real staff row.
    for e in enrollments:
        mapped = rebbe_map.get(e.pop("rebbe", ""), [])
        e["rebbe"] = mapped[0] if mapped else None

    connections = sorted(
        {(s, pid) for pid, names in conns.items() for s in names})
    (OUT / "connections.json").write_text(json.dumps(
        [dict(staff=s, person_id=pid) for s, pid in connections], indent=1))

    # Campaign touchpoints -> interactions, matched by name like the year rows.
    tp_out, tp_missed = [], 0
    for t in load_touchpoints(awb):
        f, l, nick = split_name(f"{t['first']} {t['last']}")
        hit = index.get((f, l)) or (index.get((nick, l)) if nick else None)
        if not hit:
            hit = [p for p in people if p["last"] == l and first_names_equivalent(p["first"], f)]
        if hit:
            mapped = rebbe_map.get(t["rebbe"], [])
            tp_out.append(dict(person_id=hit[0]["alumni_id"], campaign=t["campaign"],
                               staff=mapped[0] if mapped else None))
        else:
            tp_missed += 1
    (OUT / "interactions.json").write_text(json.dumps(tp_out, indent=1))
    print(f"campaign touchpoints      {len(tp_out)} matched, {tp_missed} unmatched")

    for c in cands:
        c["decision"] = prior.get(c["key"], "")

    (OUT / "people.json").write_text(json.dumps(people, indent=1))
    (OUT / "enrollments.json").write_text(json.dumps(enrollments, indent=1))
    (OUT / "merge_candidates.json").write_text(json.dumps(cands, indent=1))
    (OUT / "staff.json").write_text(json.dumps(
        dict(staff=staff, rebbe_string_map=rebbe_map, unmapped=dict(rebbe_unmapped)), indent=1))

    tiers = Counter(c["tier"] for c in cands)
    print(f"staff (rabbis)            {len(staff)}"
          f"  ({sum(1 for s in staff if s['has_own_tab'])} with a connection tab,"
          f" {len(newly_added)} found only in the alumni database)")
    if newly_added:
        print("    added: " + ", ".join(f"{n} ({c} alumni)" for n, c in newly_added))
    print(f"  'Rebbe Contact' strings {len(rebbe_strings)} distinct"
          f" -> {len(rebbe_map)} mapped, {len(rebbe_unmapped)} unmapped")
    if rebbe_unmapped:
        print("    unmapped: " + ", ".join(f"{k!r}x{v}" for k, v in rebbe_unmapped.most_common(10)))
    print(f"canonical people          {len(people)}"
          f"  ({len(merged_into)} merged away, {sum(1 for p in people if p.get('aliases'))} now carry an alias)")
    print(f"year-tab rows read        {len(year_rows)}")
    print(f"  matched to a person     {len(year_rows) - len(unmatched)}")
    print(f"  UNMATCHED               {len(unmatched)}")
    print(f"enrollments (deduped)     {len(enrollments)}")
    print(f"rebbeim connections       {sum(per_rebbe.values())} across {len(per_rebbe)} rabbis")
    print(f"merge candidates          {len(cands)}"
          f"  (certain {tiers['certain']}, likely {tiers['likely']}, review {tiers['review']})")
    if unmatched:
        print("\nfirst unmatched year-tab rows:")
        for r in unmatched[:15]:
            print(f"   {r['first']} {r['last']:<20} {r['source_tab']}")

    return people, enrollments, cands, unmatched


if __name__ == "__main__":
    main()
